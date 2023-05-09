import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.utils import column_index_from_string


class LetsCalculateUI:
    def __init__(self, window):
        self.window = window
        window.title('Lets Calculate')
        window.geometry("300x220+10+20")
        window.configure(highlightbackground="red", highlightthickness=2)

        # creating row and column labels
        tk.Label(window, text="R", padx=10, pady=10, font=('Arial', 12)).grid(row=0, column=2)
        tk.Label(window, text="C", padx=10, pady=10, font=('Arial', 12)).grid(row=0, column=4)

        # creating inpuut boxes for row and colum values
        self.row1 = tk.Entry(window, bg='lightyellow', width=6, highlightthickness=1, highlightbackground="red")
        self.row1.grid(row=2, column=2, pady=9)

        self.column1 = tk.Entry(window, bg='lightyellow', width=6, highlightthickness=1, highlightbackground="red")
        self.column1.grid(row=2, column=4, pady=9)

        self.row2 = tk.Entry(window, bg='lightyellow', width=6, highlightthickness=1, highlightbackground="red")
        self.row2.grid(row=4, column=2, pady=9)

        self.column2 = tk.Entry(window, bg='lightyellow', width=6, highlightthickness=1, highlightbackground="red")
        self.column2.grid(row=4, column=4, pady=9)

        # creating addition and substraction buttons
        self.add_button = tk.Button(window, text="+", font=('Arial', 12, 'bold'), width=4,
                                    command=lambda: self.add_sub_function("+"), fg="red", bd=1, relief='solid',
                                    highlightbackground="#42f5e6")
        self.add_button.grid(row=6, column=2, pady=9)
        self.add_button.bind("<FocusOut>", self.update_result)

        self.sub_button = tk.Button(window, text="-", font=('Arial', 12, 'bold'), width=4,
                                    command=lambda: self.add_sub_function("-"), fg="red", bd=1, relief='solid',
                                    highlightbackground="#42f5e6")
        self.sub_button.grid(row=6, column=4, pady=9)
        self.sub_button.bind("<FocusOut>", self.update_result)

        # creating Answer label and Result label to store result
        self.label1 = tk.Label(window, text='Ans:', fg='blue', font=('Arial', 14)).grid(row=8)
        self.result = tk.Label(text="", width=10, bg='lightyellow', highlightthickness=1, highlightbackground="red")
        self.result.grid(row=8, column=3, padx=4, pady=9)

    # Method to reset the result value when user changes the row,colmn value and clicks on "+" or "-"
    def update_result(self, event):
        self.result.config(text="")

    # Method contains logic to add or substract based on the given row and column values
    def add_sub_function(self, operator):

        if self.result.cget('text') != "":
            self.result.config(text="")

        result_value = self.check_read_operate(operator)
        # Seting the result value in result grid
        self.result.config(text=result_value)

    # Method to check the given row and column values, handle error and calculate
    def check_read_operate(self, operator):
        r1 = self.row1.get()
        c1 = self.column1.get()
        r2 = self.row2.get()
        c2 = self.column2.get()

        if not r1 or not c1 or not r2 or not c2:
            messagebox.showerror("Error", "Please fill all fields.")
            return
        try:
            r1, r2 = int(r1), int(r2)
            if r1 <= 0 or r2 <= 0:
                messagebox.showerror("Error", "Enter value greater than 0 for rows")
                return

        except ValueError:
            messagebox.showerror("Error", "Enter only integer values for rows")
            return

        try:
            c1, c2 = column_index_from_string(c1), column_index_from_string(c2)
        except ValueError:
            messagebox.showerror("Error", "Enter valid column letters ('A', 'B', 'C', etc.) for columns !!")
            return

        # reading the Excel file here
        try:
            workbook = openpyxl.load_workbook('Calculate.xlsx')
            sheet = workbook.active
        except:
            messagebox.showerror("Error", "Could not open Excel file or Check if the file exists !!!")
            return

        # adding logic for cell data here
        try:
            row_col_value1 = sheet.cell(row=r1, column=c1).value
            if row_col_value1 is None:
                col_letter = openpyxl.utils.get_column_letter(c1)
                messagebox.showerror("Error", f"Cell({r1},{col_letter})does not have a value")
                return
            row_col_value2 = sheet.cell(row=r2, column=c2).value
            if row_col_value2 is None:
                col_letter = openpyxl.utils.get_column_letter(c2)
                messagebox.showerror("Error", f"Cell({r2},{col_letter})does not have a value")
                return

        except:
            messagebox.showerror("Error", "Error in retrieving data from Excel file!! ")
            return

        if operator == "+":
            result_value = row_col_value1 + row_col_value2
        else:
            result_value = row_col_value1 - row_col_value2

        return result_value


window = tk.Tk()
calculateUI = LetsCalculateUI(window)
window.mainloop()